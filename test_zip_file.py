
import os.path
import os
import zipfile
from PyPDF2 import PdfReader
import xlrd
from openpyxl import load_workbook


resources_dir = "resources"
tmp_dir = "tmp"

if not os.path.exists(tmp_dir):
    os.makedirs(tmp_dir)

files_to_archive = ["file1.xls", "file2.xlsx", "file3.pdf", "file4.txt"]

with zipfile.ZipFile(os.path.join(tmp_dir, "file.zip"), "w") as zipf:
    for file_to_archive in files_to_archive:
        file_path = os.path.join(resources_dir, file_to_archive)
        zipf.write(file_path, os.path.basename(file_path))

def test_check_xls():
    # Открываем архив
    archive_path = os.path.join(tmp_dir, "file.zip")
    with zipfile.ZipFile(archive_path, "r") as zipf:
        file_to_check = "file1.xls"
        # Проверка, что файл есть в архиве
        assert file_to_check in zipf.namelist()
        # Открываем файл в архиве
        with zipf.open(file_to_check) as xls_data:
            # Открываем файл Excel XLS с использованием xlrd
            book = xlrd.open_workbook(file_contents=xls_data.read())
            sheet = book.sheet_by_index(0)
            cell_value = sheet.cell_value(2, 1)
            # Проверяем, что значение в ячейке соответствует ожидаемому значению
            assert cell_value == 'Mara'

def test_check_xlsx():
    # Открываем архив
    archive_path = os.path.join(tmp_dir, "file.zip")
    with zipfile.ZipFile(archive_path, "r") as zipf:
        file_to_check = "file2.xlsx"
        # Проверка, что файл есть в архиве
        assert file_to_check in zipf.namelist()
        # Открываем файл в архиве
        with zipf.open(file_to_check) as xlsx_data:
            # Загружаем файл XLSX с использованием openpyxl
            workbook = load_workbook(xlsx_data, read_only=True)
            sheet = workbook.active
            # Получаем значение ячейки (1, 2), где 1 - номер строки, 2 - номер столбца
            cell_value = sheet.cell(row=2, column=3).value
            # Проверяем, что значение в ячейке соответствует ожидаемому значению
            assert cell_value == 'Abril'

def test_check_pdf():
    # Открываем архив
    archive_path = os.path.join(tmp_dir, "file.zip")
    with zipfile.ZipFile(archive_path, "r") as zipf:
        file_to_check = "file3.pdf"
        # Проверка, что файл есть в архиве
        assert file_to_check in zipf.namelist()
        # Открываем файл в архиве
        with zipf.open(file_to_check) as pdf_data:
            # Открываем файл PDF с использованием PyPDF2
            pdf = PdfReader(pdf_data)
            # Получаем текст из файла
            text = ""
            for page in pdf.pages:
                text += page.extract_text()
            # Проверяем, что текст содержит ожидаемую фразу
            expected_phrase = "Python Testing with pytest"
            assert expected_phrase in text

def test_check_txt():
    # Открываем архив
    archive_path = os.path.join(tmp_dir, "file.zip")
    with zipfile.ZipFile(archive_path, "r") as zipf:
        file_to_check = "file4.txt"
        # Проверка, что файл есть в архиве
        assert file_to_check in zipf.namelist()
        # Открываем файл в архиве
        with zipf.open(file_to_check) as txt_file:
            # Читаем содержимое файла
            file_content = txt_file.read().decode("utf-8")
            # Ожидаемое содержимое файла
            expected_content = "\nText Text "
            # Проверяем, что содержимое файла соответствует ожидаемому содержимому
            assert file_content == expected_content